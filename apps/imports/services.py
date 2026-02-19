from pathlib import Path
import importlib

from django.db import transaction
from openpyxl import load_workbook

from apps.accounts.models import UserRole
from apps.core.models import Client, Notification, NotificationType

from .models import ImportBatch, ImportBatchStatus, RawExcelRow

REQUIRED_COLUMNS = {"phone", "full_name"}


def _normalize_header(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _read_xlsx(file_path: str):
    workbook = load_workbook(file_path)
    sheet = workbook.active
    header = [_normalize_header(cell.value) for cell in sheet[1]]
    rows = [
        (row_num, tuple(row_values))
        for row_num, row_values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2)
    ]
    return header, rows


def _read_xls(file_path: str):
    try:
        xlrd = importlib.import_module("xlrd")
    except ModuleNotFoundError as exc:
        raise ValueError(
            "XLS import requires xlrd package. Install xlrd or upload file in XLSX format."
        ) from exc

    workbook = xlrd.open_workbook(file_path)
    sheet = workbook.sheet_by_index(0)
    header = [_normalize_header(sheet.cell_value(0, col)) for col in range(sheet.ncols)]
    rows = []
    for row_index in range(1, sheet.nrows):
        values = tuple(sheet.cell_value(row_index, col) for col in range(sheet.ncols))
        rows.append((row_index + 1, values))
    return header, rows


def process_import_batch_sync(batch_id: int, file_path: str) -> None:
    batch = ImportBatch.objects.get(pk=batch_id)
    batch.status = ImportBatchStatus.PROCESSING
    batch.error_message = ""
    batch.save(update_fields=["status", "error_message"])

    try:
        extension = Path(file_path).suffix.lower()
        if extension == ".xls":
            header, rows = _read_xls(file_path)
        else:
            header, rows = _read_xlsx(file_path)

        header_set = {h for h in header if h}
        if not REQUIRED_COLUMNS.issubset(header_set):
            missing = REQUIRED_COLUMNS - header_set
            raise ValueError(f"Missing required columns: {', '.join(sorted(missing))}")

        indexes = {name: header.index(name) for name in header if name}
        stats = {"rows_total": 0, "clients_created": 0, "clients_updated": 0}

        def _save_row(row_num, row_values):
            stats["rows_total"] += 1
            row_data = {
                col: row_values[position] if position < len(row_values) else None
                for col, position in indexes.items()
            }
            RawExcelRow.objects.create(batch=batch, row_index=row_num, raw_data=row_data)
            phone = str(row_data.get("phone", "")).strip()
            full_name = str(row_data.get("full_name", "")).strip()
            if phone:
                _, created = Client.objects.update_or_create(phone=phone, defaults={"full_name": full_name or phone})
                if created:
                    stats["clients_created"] += 1
                else:
                    stats["clients_updated"] += 1

        with transaction.atomic():
            batch.rows.all().delete()
            for row_num, row_values in rows:
                _save_row(row_num, row_values)

        batch.status = ImportBatchStatus.DONE
        batch.error_message = ""
        batch.save(update_fields=["status", "error_message"])
        _create_import_summary_notifications(
            batch=batch,
            title=f"Импорт #{batch.id} завершен",
            body=(
                f"Файл: {batch.source_file_name or 'unknown'}; "
                f"строк: {stats['rows_total']}; "
                f"новых клиентов: {stats['clients_created']}; "
                f"обновленных клиентов: {stats['clients_updated']}."
            ),
        )
    except Exception as exc:
        batch.status = ImportBatchStatus.FAILED
        batch.error_message = str(exc)
        batch.save(update_fields=["status", "error_message"])
        _create_import_summary_notifications(
            batch=batch,
            title=f"Импорт #{batch.id} завершился ошибкой",
            body=f"Ошибка: {exc}",
        )
        raise


def _create_import_summary_notifications(batch: ImportBatch, title: str, body: str) -> None:
    recipients = {batch.uploaded_by_id}
    manager_ids = batch.uploaded_by.__class__.objects.filter(role__in=[UserRole.OWNER, UserRole.ADMIN]).values_list(
        "id",
        flat=True,
    )
    recipients.update(manager_ids)
    for recipient_id in recipients:
        Notification.objects.create(
            recipient_id=recipient_id,
            notification_type=NotificationType.IMPORT_SUMMARY,
            title=title,
            body=body,
        )
